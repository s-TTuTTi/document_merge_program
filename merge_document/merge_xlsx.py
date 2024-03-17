from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from copy import copy
import pandas as pd
from functools import reduce


def copy_sheet_data(source_sheet, target_sheet, start_row=1):

    # 만약 start_row가 1이 아니면 merged_cells를 조정
    if start_row != 1:
        for merged_cell in source_sheet.merged_cells:
            range = merged_cell.bounds

            # start_row를 더해서 새로운 행 번호 생성
            new_first_row = range[1] + start_row - 1
            new_last_row = range[3] + start_row - 1

            # 새로운 셀 좌표 문자열 생성
            new_merged_cell_range = get_column_letter(range[0]) + str(new_first_row) + ":" + get_column_letter(range[2]) + str(new_last_row)

            # target_sheet에 추가
            target_sheet.merged_cells.add(new_merged_cell_range)

            # 각 열의 넓이를 복사(두 열의 넓이 중 더 큰 것을 선택)
            for col_letter, column_dimensions in source_sheet.column_dimensions.items():
                if target_sheet.column_dimensions[col_letter].width < column_dimensions.width:
                    target_sheet.column_dimensions[col_letter].width = column_dimensions.width

            # 각 행의 높이를 복사
            for row_letter, row_dimensions in source_sheet.row_dimensions.items():
                target_sheet.row_dimensions[row_letter].height = row_dimensions.height
    else:
        # 병합된 셀 정보 복사
        target_sheet.merged_cells = source_sheet.merged_cells

        # 각 열의 넓이를 복사
        for col_letter, column_dimensions in source_sheet.column_dimensions.items():
            target_sheet.column_dimensions[col_letter].width = column_dimensions.width

        # 각 행의 높이를 복사
        for row_letter, row_dimensions in source_sheet.row_dimensions.items():
            target_sheet.row_dimensions[row_letter].height = row_dimensions.height

    # 각 셀의 데이터 및 스타일 복사
    for row in source_sheet.iter_rows(min_row=1, max_row=source_sheet.max_row, min_col=1,
                                      max_col=source_sheet.max_column):
        for cell in row:
            # 새로운 셀 생성
            new_cell = target_sheet[get_column_letter(cell.column) + str(start_row + cell.row - 1)]

            # 셀 데이터 복사
            new_cell.value = cell.value

            # 셀 스타일 복사
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)


def merge_xlsx_multi_sheet(sheet_list, output_file_name):
    # 새로운 워크북 생성
    new_wb = Workbook()
    new_wb.remove(new_wb.active)

    page_num = 1

    for file_info in sheet_list:
        new_ws = new_wb.create_sheet("Page " + str(page_num))

        file_name = file_info[0]
        sheet_names = file_info[1:]

        # 원본 워크북 로드
        wb = load_workbook(filename=file_name)

        for sheet_name in sheet_names:
            # 원본 워크시트 로드
            ws = wb[sheet_name]

            # 데이터 및 스타일 복사
            copy_sheet_data(ws, new_ws)

            # 원본 워크북 닫기
            wb.close()

        page_num += 1

    # 새로운 워크북 저장
    new_wb.save(output_file_name)
    new_wb.close()

def merge_xlsx_single_sheet(sheet_list, output_file_name):
    # 새로운 워크북 생성
    new_wb = Workbook()
    new_wb.remove(new_wb.active)
    new_ws = new_wb.create_sheet("Page 1")

    start_row = 1

    for file_info in sheet_list:
        file_name = file_info[0]
        sheet_names = file_info[1:]

        # 원본 워크북 로드
        wb = load_workbook(filename=file_name)

        for sheet_name in sheet_names:
            # 원본 워크시트 로드
            ws = wb[sheet_name]

            # 데이터 및 스타일 복사
            copy_sheet_data(ws, new_ws, start_row)

            # 원본 워크북 닫기
            wb.close()

            start_row = new_ws.max_row + 2

    # 새로운 워크북 저장
    new_wb.save(output_file_name)
    new_wb.close()

def merge_xlsx_single_table(sheet_list, output_file_name):
    df_list = []
    for file_info in sheet_list:
        file_name = file_info[0]
        sheet_names = file_info[1:]

        for sheet_name in sheet_names:
            df = pd.read_excel(file_name, sheet_name=sheet_name, engine='openpyxl')
            df_list.append(df)

    merged_df = reduce(lambda left, right: pd.merge(left, right, how='outer'), df_list)
    print(merged_df)
    print(merged_df.columns)
    merged_df.to_excel(output_file_name, sheet_name='Page1', index=False, engine='openpyxl')





    column_widths = {}
    title_cell = None
    data_cell = None
    title_row_height = 0
    data_row_height = 0

    for file_info in sheet_list:
        file_name = file_info[0]
        sheet_names = file_info[1:]

        wb = load_workbook(filename=file_name)

        for sheet_name in sheet_names:
            ws = wb[sheet_name]

            if title_cell is None or data_cell is None:
                title_cell = ws.cell(row=1, column=1)
                data_cell = ws.cell(row=2, column=1)
            if title_row_height < ws.row_dimensions[1].height:
                title_row_height = ws.row_dimensions[1].height
            if data_row_height < ws.row_dimensions[2].height:
                data_row_height = ws.row_dimensions[2].height

            for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if cell.value in merged_df.columns:
                        current_width = ws.column_dimensions[cell.column_letter].width
                        # 열 제목이 딕셔너리에 이미 존재하고, 현재 넓이가 저장된 넓이보다 큰 경우에만 업데이트
                        if cell.value not in column_widths or current_width > column_widths[cell.value]:
                            column_widths[cell.value] = current_width

                print(column_widths)  # 저장된 열 제목과 넓이 출력





    new_wb = load_workbook(filename=output_file_name)
    new_ws = new_wb['Page1']

    for row in new_ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=new_ws.max_column):
        for cell in row:
            # 셀 스타일 복사
            cell.font = copy(title_cell.font)
            cell.border = copy(title_cell.border)
            cell.fill = copy(title_cell.fill)
            cell.number_format = copy(title_cell.number_format)
            cell.protection = copy(title_cell.protection)
            cell.alignment = copy(title_cell.alignment)

    for row in new_ws.iter_rows(min_row=2, max_row=new_ws.max_row, min_col=1, max_col=new_ws.max_column):
        for cell in row:
            # 셀 스타일 복사
            cell.font = copy(data_cell.font)
            cell.border = copy(data_cell.border)
            cell.fill = copy(data_cell.fill)
            cell.number_format = copy(data_cell.number_format)
            cell.protection = copy(data_cell.protection)
            cell.alignment = copy(data_cell.alignment)

    # 열 너비 설정
    for col in new_ws.columns:
        column = col[0].column_letter
        if col[0].value in column_widths:
            new_ws.column_dimensions[column].width = column_widths[col[0].value]

    # 행 높이 설정
    new_ws.row_dimensions[1].height = title_row_height
    for row in range(2, new_ws.max_row + 1):
        new_ws.row_dimensions[row].height = data_row_height

    # 변경사항 저장
    new_wb.save(output_file_name)