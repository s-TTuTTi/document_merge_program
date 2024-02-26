from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string
from copy import copy


def copy_sheet_data(source_sheet, target_sheet, start_row=1):

    # 만약 start_row가 1이 아니면 merged_cells를 조정
    if start_row != 1:
        for merged_cell in source_sheet.merged_cells:
            range = merged_cell.bounds

            # start_row를 더해서 새로운 행 번호 생성
            new_first_row = range[1] + start_row - 1
            new_last_row = range[3] + start_row - 1

            # 새로운 셀 좌표 문자열 생성
            new_merged_cell_range = get_column_letter(range[0]) + str(new_first_row) + ":" + get_column_letter(range[2]) + str(new_last_row)

            # target_sheet에 추가
            target_sheet.merged_cells.add(new_merged_cell_range)

            # 각 열의 넓이를 복사
            for col_letter, column_dimensions in source_sheet.column_dimensions.items():
                if target_sheet.column_dimensions[col_letter].width < column_dimensions.width:
                    target_sheet.column_dimensions[col_letter].width = column_dimensions.width

            # 각 행의 높이를 복사
            for row_letter, row_dimensions in source_sheet.row_dimensions.items():
                target_sheet.row_dimensions[row_letter].height = row_dimensions.height
    else:
        # 병합된 셀 정보 복사
        target_sheet.merged_cells = source_sheet.merged_cells

        # 각 열의 넓이를 복사
        for col_letter, column_dimensions in source_sheet.column_dimensions.items():
            target_sheet.column_dimensions[col_letter].width = column_dimensions.width

        # 각 행의 높이를 복사
        for row_letter, row_dimensions in source_sheet.row_dimensions.items():
            target_sheet.row_dimensions[row_letter].height = row_dimensions.height

    # 각 셀의 데이터 및 스타일 복사
    for row in source_sheet.iter_rows(min_row=1, max_row=source_sheet.max_row, min_col=1,
                                      max_col=source_sheet.max_column):
        for cell in row:
            # 새로운 셀 생성
            new_cell = target_sheet[get_column_letter(cell.column) + str(start_row + cell.row - 1)]

            # 셀 데이터 복사
            new_cell.value = cell.value

            # 셀 스타일 복사
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)


def merge_xlsx_multi_sheet():
    # 원본 워크북 로드
    wb = load_workbook(filename="xlsx_sample/test2.xlsx")
    ws = wb["참여학사조직"]

    twb = load_workbook(filename="xlsx_sample/test1.xlsx")
    tws = twb["창원대"]

    # 새로운 워크북 생성
    newwb = Workbook()
    newwb.remove(newwb.active)
    newws1 = newwb.create_sheet("Page 1")
    newws2 = newwb.create_sheet("Page 2")

    # 데이터 및 스타일 복사
    copy_sheet_data(ws, newws1)
    copy_sheet_data(tws, newws2)

    # 새로운 워크북 저장
    newwb.save("xlsx_sample/output_multi_sheet.xlsx")

    # 원본 워크북과 새로운 워크북 닫기
    wb.close()
    newwb.close()

def merge_xlsx_single_sheet():
    # 원본 워크북 로드
    wb = load_workbook(filename="xlsx_sample/test2.xlsx")
    ws = wb["참여학사조직"]

    twb = load_workbook(filename="xlsx_sample/test1.xlsx")
    tws = twb["창원대"]

    # 새로운 워크북 생성
    newwb = Workbook()
    newwb.remove(newwb.active)
    newws = newwb.create_sheet("Page 1")

    # 데이터 및 스타일 복사
    copy_sheet_data(ws, newws)
    copy_sheet_data(tws, newws, newws.max_row + 1)

    # 새로운 워크북 저장
    newwb.save("xlsx_sample/output_single_sheet.xlsx")

    # 원본 워크북과 새로운 워크북 닫기
    wb.close()
    newwb.close()

merge_xlsx_multi_sheet()
merge_xlsx_single_sheet()
