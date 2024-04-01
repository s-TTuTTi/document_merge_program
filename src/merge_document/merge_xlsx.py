from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from copy import copy
import pandas as pd
from functools import reduce

class ExcelSheetHandler:
    def __init__(self, sheet):
        self.sheet = sheet

    def identify_table(self, engine='pandas'):
        if engine == 'pandas':
            return self.identify_table_pandas()
        elif engine == 'openpyxl':
            return self.identify_table_openpyxl()
        else:
            raise ValueError("Invalid engine. Supported engines are 'openpyxl' and 'pandas'.")

    def identify_table_pandas(self):
        for i in range(len(self.sheet)):
            for j in range(len(self.sheet.columns)):
                if self.check_table_pandas(i, j):
                    return i + 1, j + 1

    def identify_table_openpyxl(self):
        for row in self.sheet.iter_rows(min_row=1, min_col=1):
            for cell in row:
                if cell.value is not None:
                    if self.check_table_openpyxl(cell):
                        return cell.row, cell.column

    def check_table_pandas(self, row, col):
        if self.sheet.iloc[row, col]:  # 데이터가 있는 셀일 경우
            # 3행 3열까지 데이터가 있는 지 검사
            return (self.sheet.iloc[row:row + 3, col:col + 3].notnull().sum().sum()) == 9

    def check_table_openpyxl(self, cell):
        # 해당 셀을 기준으로 3행 3열까지 데이터가 있는 지 검사
        return all(self.sheet.cell(row=cell.row + i, column=cell.column + j).value is not None
                   for i in range(3) for j in range(3))

    @staticmethod
    def move_cell_range(cell, row=1, column=1):  # row로 몇칸 column으로 몇칸 이동한 셀 범위를 리턴
        first_column, first_row, last_column, last_row = cell.bounds  # (열,행,열,행)

        new_first_column = first_column + column - 1
        new_first_row = first_row + row - 1

        new_last_column = last_column + column - 1
        new_last_row = last_row + row - 1

        new_cell_range = f"{get_column_letter(new_first_column)}{new_first_row}:{get_column_letter(new_last_column)}{new_last_row}"

        return new_cell_range

    @staticmethod
    def copy_cell_style(new_cell, cell):
        new_cell.font = copy(cell.font)
        new_cell.border = copy(cell.border)
        new_cell.fill = copy(cell.fill)
        new_cell.number_format = copy(cell.number_format)
        new_cell.protection = copy(cell.protection)
        new_cell.alignment = copy(cell.alignment)

    def copy_merged_cell(self, target_sheet, start_row=1, start_column=1):
        for merged_cell in self.sheet.merged_cells:
            new_merged_cell_range = self.move_cell_range(cell=merged_cell, row=start_row, column=start_column)
            target_sheet.merged_cells.add(new_merged_cell_range)

    def copy_column_dimensions(self, target_sheet, start_column=1):
        for col_letter, column_dimensions in self.sheet.column_dimensions.items():
            col_index = column_index_from_string(col_letter)
            target_col_letter = get_column_letter(col_index + start_column - 1)

            target_column_width = target_sheet.column_dimensions[target_col_letter].width
            source_column_width = column_dimensions.width

            if source_column_width is None:
                continue

            if target_column_width is None or (target_column_width is not None and target_column_width < source_column_width):
                target_sheet.column_dimensions[target_col_letter].width = source_column_width

    def copy_row_dimensions(self, target_sheet, start_row=1):
        for row_letter, row_dimensions in self.sheet.row_dimensions.items():
            target_row_height = target_sheet.row_dimensions[row_letter + start_row - 1].height
            source_row_height = row_dimensions.height

            if source_row_height is None:
                continue

            if target_row_height is None or (target_row_height is not None and target_row_height < source_row_height):
                target_sheet.row_dimensions[row_letter + start_row - 1].height = source_row_height

    def copy_cell_value_and_style(self, target_sheet, start_row, start_column):
        for row in self.sheet.iter_rows(min_row=self.sheet.min_row, max_row=self.sheet.max_row,
                                        min_col=self.sheet.min_column, max_col=self.sheet.max_column):
            for cell in row:
                new_cell = target_sheet[get_column_letter(cell.column + start_column - 1) + str(cell.row + start_row - 1)]
                new_cell.value = cell.value
                self.copy_cell_style(new_cell, cell)

    def copy_sheet_data(self, target_sheet, start_row=1, start_column=1):  # target_sheet의 정확히 어디에 카피를 할 건지
        # 병합된 셀 처리
        self.copy_merged_cell(target_sheet, start_row, start_column)
        # 각 열의 넓이 처리
        self.copy_column_dimensions(target_sheet, start_column)
        # 각 행의 높이 처리
        self.copy_row_dimensions(target_sheet, start_row)
        # 셀 값 및 스타일 복사
        self.copy_cell_value_and_style(target_sheet, start_row, start_column)


class ExcelMerger:
    def __init__(self):
        self.new_wb = Workbook()
        self.new_wb.remove(self.new_wb.active)

    def merge_multi_sheet(self, sheet_list, output_file_name):
        page_num = 1

        for file_info in sheet_list:
            file_name = file_info[0]
            sheet_names = file_info[1:]

            # 원본 워크북 로드
            wb = load_workbook(filename=file_name)

            for sheet_name in sheet_names:
                # 원본 워크시트 로드
                ws = wb[sheet_name]
                # 새 워크시트 생성
                new_ws = self.new_wb.create_sheet("Page " + str(page_num))

                # 데이터 및 스타일 복사
                handler = ExcelSheetHandler(ws)
                handler.copy_sheet_data(new_ws)

                page_num += 1

            wb.close()

        # 새 워크북 저장
        self.save(output_file_name)

    def merge_single_sheet(self, sheet_list, output_file_name):
        # 새 워크시트 생성
        new_ws = self.new_wb.create_sheet("Page 1")

        start_row = 1
        start_column = 1

        for file_info in sheet_list:
            file_name = file_info[0]
            sheet_names = file_info[1:]

            # 원본 워크북 로드
            wb = load_workbook(filename=file_name)

            for sheet_name in sheet_names:
                # 원본 워크시트 로드
                ws = wb[sheet_name]

                # 데이터 및 스타일 복사
                handler = ExcelSheetHandler(ws)
                handler.copy_sheet_data(target_sheet=new_ws, start_row=start_row, start_column=start_column)

                start_row = new_ws.max_row + 2

            wb.close()

        # 새 워크북 저장
        self.save(output_file_name)

    def merge_single_table(self, sheet_list, output_file_name):
        df_list = []

        for file_info in sheet_list:
            file_name = file_info[0]
            sheet_names = file_info[1:]

            for sheet_name in sheet_names:
                df = pd.read_excel(file_name, sheet_name=sheet_name, engine='openpyxl', header=None)
                df_list.append(df)

        merged_df = reduce(lambda left, right: pd.merge(left, right, how='outer'), df_list)
        merged_df.to_excel(output_file_name, sheet_name='Page1', index=False, engine='openpyxl')





        column_widths = {}
        title_cell = None
        data_cell = None
        title_row_height = 0
        data_row_height = 0

        for file_info in sheet_list:
            file_name = file_info[0]
            sheet_names = file_info[1:]

            wb = load_workbook(filename=file_name)

            for sheet_name in sheet_names:
                ws = wb[sheet_name]

                if title_cell is None or data_cell is None:
                    title_cell = ws.cell(row=1, column=1)
                    data_cell = ws.cell(row=2, column=1)
                if title_row_height < ws.row_dimensions[1].height:
                    title_row_height = ws.row_dimensions[1].height
                if data_row_height < ws.row_dimensions[2].height:
                    data_row_height = ws.row_dimensions[2].height

                for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        if cell.value in merged_df.columns:
                            current_width = ws.column_dimensions[cell.column_letter].width
                            # 열 제목이 딕셔너리에 이미 존재하고, 현재 넓이가 저장된 넓이보다 큰 경우에만 업데이트
                            if cell.value not in column_widths or current_width > column_widths[cell.value]:
                                column_widths[cell.value] = current_width

                    print(column_widths)  # 저장된 열 제목과 넓이 출력





        new_wb = load_workbook(filename=output_file_name)
        new_ws = new_wb['Page1']

        for row in new_ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=new_ws.max_column):
            for cell in row:
                # 셀 스타일 복사
                cell.font = copy(title_cell.font)
                cell.border = copy(title_cell.border)
                cell.fill = copy(title_cell.fill)
                cell.number_format = copy(title_cell.number_format)
                cell.protection = copy(title_cell.protection)
                cell.alignment = copy(title_cell.alignment)

        for row in new_ws.iter_rows(min_row=2, max_row=new_ws.max_row, min_col=1, max_col=new_ws.max_column):
            for cell in row:
                # 셀 스타일 복사
                cell.font = copy(data_cell.font)
                cell.border = copy(data_cell.border)
                cell.fill = copy(data_cell.fill)
                cell.number_format = copy(data_cell.number_format)
                cell.protection = copy(data_cell.protection)
                cell.alignment = copy(data_cell.alignment)

        # 열 너비 설정
        for col in new_ws.columns:
            column = col[0].column_letter
            if col[0].value in column_widths:
                new_ws.column_dimensions[column].width = column_widths[col[0].value]

        # 행 높이 설정
        new_ws.row_dimensions[1].height = title_row_height
        for row in range(2, new_ws.max_row + 1):
            new_ws.row_dimensions[row].height = data_row_height

        # 변경사항 저장
        new_wb.save(output_file_name)

    def save(self, output_file_name):
        self.new_wb.save(output_file_name)
        self.new_wb.close()


if __name__ == "__main__":
    merger = ExcelMerger()
    sheet_list = [["xlsx_sample/test1.xlsx", "Sheet4", "Sheet2"]]
    output_file_name = "xlsx_sample/output.xlsx"
    # merger.merge_single_sheet(sheet_list, output_file_name)
    merger.merge_single_table(sheet_list, output_file_name)